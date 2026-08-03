
import json
import math
import os
import random
import re
import smtplib
import sys
import threading
import time
from datetime import date, datetime, time as clock_time, timedelta
from email.message import EmailMessage
from pathlib import Path

import numpy as np
import sounddevice as sd
from PyQt6.QtCore import QRect, QTimer, Qt
from PyQt6.QtGui import QColor, QFont, QPainter, QPen, QPixmap
from PyQt6.QtWidgets import (
    QApplication,
    QAbstractItemView,
    QFormLayout,
    QFrame,
    QGridLayout,
    QHBoxLayout,
    QLabel,
    QLineEdit,
    QMainWindow,
    QMessageBox,
    QPushButton,
    QTableWidget,
    QTableWidgetItem,
    QStackedWidget,
    QVBoxLayout,
    QWidget,
)


# In de .exe staan assets tijdelijk uitgepakt, maar gegevens moeten naast de .exe blijven.
if getattr(sys, "frozen", False):
    RESOURCE_DIR = Path(getattr(sys, "_MEIPASS", Path(sys.executable).resolve().parent))
    DATA_DIR = Path(sys.executable).resolve().parent
else:
    RESOURCE_DIR = Path(__file__).resolve().parent
    DATA_DIR = RESOURCE_DIR

HIGHSCORES_FILE = DATA_DIR / "highscores.json"
EXPORT_DIR = DATA_DIR / "daily_exports"
LOGO_FILE = RESOURCE_DIR / "logo-clean.png"
EMAIL_PATTERN = re.compile(r"^[^\s@]+@[^\s@]+\.[^\s@]+$")
NOISE_FLOOR_DBFS = -90.0
DISPLAY_MAX_DB = 132.0
SHURE_MV7_SAMPLE_RATE = 48_000
# Festival mode: require the Shure MV7 USB microphone.
REQUIRE_SHURE_MV7 = True
# Festival mode: export the finished daily leaderboard at local midnight.
DAILY_EXPORT_TIME = clock_time.min
TEST_EXPORT_SCHEDULE = False
# Configured recipient for the daily Excel export. SMTP credentials are still required to send email.
EXPORT_RECIPIENT_EMAIL = "bjornguiot@gmail.com"
SMTP_SENDER_EMAIL = "screamchallenge@gmail.com"
SMTP_HOST = "smtp.gmail.com"
SMTP_PORT = 465
SMTP_APP_PASSWORD_ENV = "MONSTER_SMTP_APP_PASSWORD"
CALIBRATION_OFFSET_DB = 141.0


class AudioLevelReader:

    def __init__(self) -> None:
        self._lock = threading.Lock()
        self._dbfs = NOISE_FLOOR_DBFS
        self._stream: sd.InputStream | None = None
        self.device_name = ""

    @staticmethod
    def _find_mv7_input() -> tuple[int | None, str]:
        for index, device in enumerate(sd.query_devices()):
            name = str(device["name"])
            if "mv7" in name.lower() and device["max_input_channels"] >= 1:
                return index, name
        if REQUIRE_SHURE_MV7:
            raise RuntimeError("Shure MV7 not found. Connect the microphone via USB and restart the app.")
        default_input = sd.query_devices(kind="input")
        return None, f"Default microphone: {default_input['name']}"

    def _audio_callback(self, indata, frames, time_info, status) -> None:
        if status:
            pass
        rms = float(np.sqrt(np.mean(np.square(indata, dtype=np.float64))))
        dbfs = 20 * math.log10(max(rms, 1e-9))
        with self._lock:
            self._dbfs = max(NOISE_FLOOR_DBFS, min(0.0, dbfs))

    def start(self) -> None:
        device_index, self.device_name = self._find_mv7_input()
        sample_rate = SHURE_MV7_SAMPLE_RATE if device_index is not None else None
        self._stream = sd.InputStream(
            device=device_index,
            channels=1,
            samplerate=sample_rate,
            blocksize=1024,
            dtype="float32",
            latency="low",
            callback=self._audio_callback,
        )
        self._stream.start()

    def stop(self) -> None:
        if self._stream is not None:
            self._stream.stop()
            self._stream.close()
            self._stream = None

    def level(self) -> float:
        with self._lock:
            return max(0.0, min(DISPLAY_MAX_DB, self._dbfs + CALIBRATION_OFFSET_DB))


class LogoMeter(QWidget):

    def __init__(self, parent=None) -> None:
        super().__init__(parent)
        self.setMinimumHeight(245)
        self._target = 0.0
        self._shown = 0.0
        self._logo = QPixmap(str(LOGO_FILE))

    def set_level(self, decibels: float) -> None:
        self._target = max(0.0, min(1.0, decibels / DISPLAY_MAX_DB))
        self._shown += (self._target - self._shown) * 0.28
        self.update()

    def paintEvent(self, event) -> None:
        painter = QPainter(self)
        painter.setRenderHint(QPainter.RenderHint.Antialiasing)
        if self._logo.isNull():
            painter.setPen(QColor("#3CD070"))
            painter.drawText(self.rect(), Qt.AlignmentFlag.AlignCenter, "Logo not found")
            return
        available = self.rect().adjusted(12, 6, -12, -6)
        ratio = self._logo.width() / self._logo.height()
        height = min(available.height(), int(available.width() / ratio))
        width = int(height * ratio)
        x = available.x() + (available.width() - width) // 2
        y = available.y() + (available.height() - height) // 2
        target = QRect(x, y, width, height)

        # The dimmed logo shows the complete meter; the bright logo grows upward.
        painter.setOpacity(0.16)
        painter.drawPixmap(target, self._logo)
        painter.save()
        filled_height = int(target.height() * self._shown)
        painter.setClipRect(target.x(), target.bottom() - filled_height + 1, target.width(), filled_height)
        painter.setOpacity(1.0)
        painter.drawPixmap(target, self._logo)
        painter.restore()


class ResultOverlay(QWidget):

    def __init__(self, parent: QWidget) -> None:
        super().__init__(parent)
        self.setAttribute(Qt.WidgetAttribute.WA_TransparentForMouseEvents)
        self.setVisible(False)
        self._timer = QTimer(self)
        self._timer.setInterval(16)
        self._timer.timeout.connect(self._advance)
        self._started_at = 0.0
        self._score = 0.0
        self._is_champion = False
        self._particles: list[dict] = []
        self._logo = QPixmap(str(LOGO_FILE))

    def show_result(self, score: float, is_champion: bool) -> None:

        self.setGeometry(self.parentWidget().rect())
        self._score = score
        self._is_champion = is_champion
        self._started_at = time.monotonic()
        self._particles = []
        if is_champion:
            for cannon_x, direction in ((self.width() * 0.10, 1), (self.width() * 0.90, -1)):
                for _ in range(70):
                    self._particles.append({
                        "x": cannon_x,
                        "y": self.height() - 95,
                        "vx": direction * random.uniform(260, 680),
                        "vy": random.uniform(-900, -390),
                        "size": random.randint(24, 52),
                        "rotation": random.randint(0, 359),
                        "spin": random.uniform(-420, 420),
                    })
        self.show()
        self.raise_()
        self._timer.start()
        self.update()

    def _advance(self) -> None:

        elapsed = time.monotonic() - self._started_at
        if elapsed > 3.0:
            self._timer.stop()
            self.hide()
            return
        for particle in self._particles:
            particle["x"] += particle["vx"] * 0.016
            particle["y"] += particle["vy"] * 0.016
            particle["vy"] += 850 * 0.016
            particle["rotation"] += particle["spin"] * 0.016
        self.update()

    @staticmethod
    def _draw_cannon(painter: QPainter, x: float, y: float, direction: int) -> None:
        """Draws a highly visible Monster-style confetti cannon pointing at the centre."""
        painter.save()
        painter.translate(x, y)
        painter.setPen(QPen(QColor("#050505"), 5))
        painter.setBrush(QColor("#202020"))
        painter.drawEllipse(-43, -10, 86, 31)
        painter.setBrush(QColor("#3CD070"))
        painter.drawEllipse(-25, -20, 50, 20)

        painter.scale(direction, 1)
        painter.rotate(-31)
        painter.setBrush(QColor("#153B25"))
        painter.drawRoundedRect(0, -20, 132, 40, 12, 12)
        painter.setBrush(QColor("#3CD070"))
        painter.drawRoundedRect(8, -13, 112, 26, 9, 9)
        painter.setBrush(QColor("#08140D"))
        painter.drawEllipse(109, -25, 39, 50)
        painter.setBrush(QColor("#E6FF00"))
        painter.drawEllipse(121, -14, 27, 28)
        painter.setPen(QPen(QColor("#E6FF00"), 4))
        for offset in (-20, 0, 20):
            painter.drawLine(148, 0, 170, offset)
        painter.restore()

    def paintEvent(self, event) -> None:
        painter = QPainter(self)
        painter.setRenderHint(QPainter.RenderHint.Antialiasing)
        painter.fillRect(self.rect(), QColor(0, 0, 0, 135))

        title = "NEW #1!" if self._is_champion else "TEST COMPLETE!"
        painter.setPen(QColor("#3CD070"))
        painter.setFont(QFont("Arial", 42, QFont.Weight.Black))
        painter.drawText(self.rect().adjusted(0, 105, 0, -120), Qt.AlignmentFlag.AlignHCenter, title)
        painter.setPen(QColor("#FFFFFF"))
        painter.setFont(QFont("Arial", 24, QFont.Weight.Bold))
        painter.drawText(self.rect().adjusted(0, 190, 0, -60), Qt.AlignmentFlag.AlignHCenter, f"{self._score:.1f} dB SPL")

        if not self._is_champion:
            return
        self._draw_cannon(painter, 85, self.height() - 62, 1)
        self._draw_cannon(painter, self.width() - 85, self.height() - 62, -1)
        for particle in self._particles:
            painter.save()
            painter.translate(particle["x"], particle["y"])
            painter.rotate(particle["rotation"])
            size = particle["size"]
            painter.drawPixmap(QRect(-size // 2, -size // 2, size, size), self._logo)
            painter.restore()


class TVRankingWindow(QMainWindow):

    def __init__(self) -> None:
        super().__init__()
        self.setWindowTitle("Monster Energy — Live Ranking")
        self.setStyleSheet("""
            QMainWindow, QWidget { background: #111111; color: #f5f5f5; }
            QTableWidget { background: #181818; color: white; border: 1px solid #3CD070;
                           font-size: 22px; gridline-color: #333; }
            QHeaderView::section { background: #242424; color: #3CD070; font-size: 18px;
                                   padding: 14px; border: none; font-weight: bold; }
        """)
        root = QWidget()
        self.setCentralWidget(root)
        self.root = root
        layout = QVBoxLayout(root)
        layout.setContentsMargins(50, 38, 50, 48)
        self.pages = QStackedWidget()
        layout.addWidget(self.pages)

        self.test_page = QWidget()
        test_layout = QVBoxLayout(self.test_page)
        test_layout.setSpacing(18)
        test_title = QLabel("MONSTER ENERGY SCREAM CHALLENGE")
        test_title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        test_title.setFont(QFont("Arial", 32, QFont.Weight.Black))
        test_title.setStyleSheet("color: #3CD070;")
        test_layout.addWidget(test_title)
        self.test_db_label = QLabel("0.0 dB SPL")
        self.test_db_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.test_db_label.setFont(QFont("Arial", 54, QFont.Weight.Black))
        self.test_db_label.setStyleSheet("color: #3CD070;")
        test_layout.addWidget(self.test_db_label)
        self.test_status_label = QLabel("GET READY FOR THE SCREAM TEST")
        self.test_status_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.test_status_label.setFont(QFont("Arial", 20, QFont.Weight.Bold))
        test_layout.addWidget(self.test_status_label)
        self.test_meter = LogoMeter()
        self.test_meter.setMinimumHeight(420)
        test_layout.addWidget(self.test_meter, 1)
        self.pages.addWidget(self.test_page)

        self.leaderboard_page = QWidget()
        leaderboard_layout = QVBoxLayout(self.leaderboard_page)
        leaderboard_layout.setSpacing(16)
        title_row = QHBoxLayout()
        title_row.addWidget(self._logo_label(68))
        title = QLabel("MONSTER ENERGY SCREAM CHALLENGE — LIVE TOP 10")
        title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        title.setFont(QFont("Arial", 28, QFont.Weight.Black))
        title.setStyleSheet("color: #3CD070;")
        title_row.addWidget(title, 1)
        title_row.addWidget(self._logo_label(68))
        leaderboard_layout.addLayout(title_row)
        self.table = QTableWidget(0, 3)
        self.table.setHorizontalHeaderLabels(["RANK", "NAME", "MAX SCORE (dB SPL)"])
        self.table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.table.setSelectionMode(QAbstractItemView.SelectionMode.NoSelection)
        self.table.verticalHeader().setVisible(False)
        self.table.setColumnWidth(0, 140)
        self.table.setColumnWidth(1, 360)
        self.table.horizontalHeader().setStretchLastSection(True)
        leaderboard_layout.addWidget(self.table, 1)
        self.pages.addWidget(self.leaderboard_page)
        self.result_overlay = ResultOverlay(root)
        self.leaderboard_timer = QTimer(self)
        self.leaderboard_timer.setSingleShot(True)
        self.leaderboard_timer.timeout.connect(self.show_leaderboard)

    @staticmethod
    def _logo_label(size: int) -> QLabel:
        logo = QLabel()
        logo.setPixmap(QPixmap(str(LOGO_FILE)))
        logo.setFixedSize(size, size)
        logo.setScaledContents(True)
        return logo

    def set_scores(self, scores: list[dict]) -> None:

        top_ten = scores[:10]
        self.table.setRowCount(len(top_ten))
        for row, score in enumerate(top_ten):
            full_name = str(score.get("name", "Unknown")).strip()
            first_name = full_name.split()[0] if full_name else "Unknown"
            values = [str(row + 1), first_name, f"{float(score.get('max_db', 0)):.1f}"]
            for column, value in enumerate(values):
                item = QTableWidgetItem(value)
                item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                item.setFont(QFont("Arial", 22, QFont.Weight.Bold if column != 1 else QFont.Weight.Normal))
                self.table.setItem(row, column, item)
            self.table.setRowHeight(row, 54)

    def show_on_tv(self) -> bool:

        screens = QApplication.screens()
        if len(screens) < 2:
            return False
        target_screen = screens[1]
        self.setGeometry(target_screen.geometry())
        self.showFullScreen()
        self.show_test(0.0, "GET READY FOR THE SCREAM TEST")
        return True

    def show_test(self, decibels: float, status: str) -> None:
        """Displays the live scream test first on the external TV."""
        self.pages.setCurrentWidget(self.test_page)
        self.test_db_label.setText(f"{decibels:.1f} dB SPL")
        self.test_status_label.setText(status)
        self.test_meter.set_level(decibels)

    def update_test(self, decibels: float, remaining: float) -> None:
        """Updates the large TV test display while the participant is screaming."""
        if self.isVisible():
            self.show_test(decibels, f"SCREAM NOW — {remaining:.1f} SECONDS LEFT")

    def show_leaderboard(self) -> None:
        """Switches the TV from the completed test to the updated leaderboard."""
        self.pages.setCurrentWidget(self.leaderboard_page)

    def show_result(self, score: float, is_champion: bool) -> None:
        """Shows the completed test and champion confetti on the external TV."""
        if self.isVisible():
            self.result_overlay.show_result(score, is_champion)
            self.leaderboard_timer.start(3000)

    def resizeEvent(self, event) -> None:
        """Keeps the result animation full-screen when the TV resolution changes."""
        super().resizeEvent(event)
        if hasattr(self, "result_overlay"):
            self.result_overlay.setGeometry(self.centralWidget().rect())


class SoundboardWindow(QMainWindow):

    def __init__(self) -> None:
        super().__init__()
        self.setWindowTitle("Monster Energy Scream Challenge")
        self.setMinimumSize(1000, 620)
        self.resize(1180, 720)
        self.audio = AudioLevelReader()
        self.measurement_timer = QTimer(self)
        self.measurement_timer.setInterval(40)
        self.measurement_timer.timeout.connect(self.update_measurement)
        self.daily_export_timer = QTimer(self)
        self.daily_export_timer.setSingleShot(True)
        self.daily_export_timer.timeout.connect(self.export_at_end_of_day)
        self.started_at = 0.0
        self.max_db = 0.0
        self.is_measuring = False
        self.ranking_display = TVRankingWindow()
        self._build_ui()
        self.load_highscores()
        self.schedule_daily_export()

    def _build_ui(self) -> None:
        self.setStyleSheet("""
            QMainWindow, QWidget { background: #111111; color: #f0f0f0; }
            QLineEdit { background: #1e1e1e; border: 1px solid #444; border-radius: 5px;
                        padding: 10px; color: white; font-size: 14px; }
            QLineEdit:focus { border: 1px solid #3CD070; }
            QPushButton { background: #3CD070; color: #07160c; border: none; border-radius: 6px;
                          padding: 12px 18px; font-weight: bold; font-size: 15px; }
            QPushButton:hover { background: #54e786; }
            QPushButton:disabled { background: #2b573c; color: #94a79a; }
            QTableWidget { background: #181818; gridline-color: #333; border: 1px solid #333;
                           border-radius: 6px; color: #eee; }
            QHeaderView::section { background: #242424; color: #3CD070; padding: 8px;
                                   border: none; font-weight: bold; }
        """)
        root = QWidget()
        self.setCentralWidget(root)
        layout = QVBoxLayout(root)
        layout.setContentsMargins(28, 24, 28, 28)
        layout.setSpacing(20)

        header_panel = QFrame()
        header_panel.setMinimumHeight(78)
        header_panel.setStyleSheet("QFrame { background: #181818; border: 1px solid #306b45; border-radius: 8px; }")
        header_row = QHBoxLayout(header_panel)
        header_row.setContentsMargins(22, 10, 22, 10)
        header_row.setSpacing(14)
        header_row.addStretch(1)
        header_row.addWidget(self._make_logo_label(header_panel, 48))
        header = QLabel("MONSTER ENERGY SCREAM CHALLENGE")
        header.setAlignment(Qt.AlignmentFlag.AlignCenter)
        header.setFont(QFont("Arial", 26, QFont.Weight.Black))
        header.setStyleSheet("background: transparent; color: #3CD070; border: none; letter-spacing: 2px;")
        header_row.addWidget(header)
        header_row.addWidget(self._make_logo_label(header_panel, 48))
        header_row.addStretch(1)
        layout.addWidget(header_panel)

        content = QHBoxLayout()
        content.setSpacing(24)
        layout.addLayout(content, 1)

        left = QVBoxLayout()
        left.setSpacing(16)
        content.addLayout(left, 3)

        registration = QFrame()
        registration.setStyleSheet("QFrame { background: #181818; border: 1px solid #303030; border-radius: 8px; }")
        form = QFormLayout(registration)
        form.setContentsMargins(18, 16, 18, 16)
        form.setSpacing(12)
        self.name_input = QLineEdit()
        self.name_input.setPlaceholderText("First and last name")
        self.email_input = QLineEdit()
        self.email_input.setPlaceholderText("name@example.com")
        self.name_input.textChanged.connect(self.validate_form)
        self.email_input.textChanged.connect(self.validate_form)
        form.addRow("Full name", self.name_input)
        form.addRow("Email address", self.email_input)
        left.addWidget(registration)

        self.db_label = QLabel("0.0 dB")
        self.db_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.db_label.setFont(QFont("Arial", 43, QFont.Weight.Bold))
        self.db_label.setStyleSheet("color: #3CD070; margin-top: 10px;")
        left.addWidget(self.db_label)
        self.status_label = QLabel("Connect the Shure MV7 via USB, then enter your details.")
        self.status_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.status_label.setStyleSheet("color: #bdbdbd; font-size: 14px;")
        left.addWidget(self.status_label)
        export_note = QLabel("Daily Top 10 export: daily_exports")
        export_note.setAlignment(Qt.AlignmentFlag.AlignCenter)
        export_note.setStyleSheet("color: #777; font-size: 11px;")
        left.addWidget(export_note)
        self.meter = LogoMeter()
        left.addWidget(self.meter)
        self.start_button = QPushButton("START SCREAM TEST")
        self.start_button.setMinimumHeight(54)
        self.start_button.setEnabled(False)
        self.start_button.clicked.connect(self.start_test)
        left.addWidget(self.start_button)
        left.addStretch()

        ranking_box = QFrame()
        ranking_box.setStyleSheet("QFrame { background: #181818; border: 1px solid #303030; border-radius: 8px; }")
        ranking_layout = QVBoxLayout(ranking_box)
        ranking_title = QLabel("TOP 10 LEADERBOARD")
        ranking_title.setFont(QFont("Arial", 15, QFont.Weight.Bold))
        ranking_title.setStyleSheet("color: #3CD070; border: none;")
        ranking_layout.addWidget(ranking_title)
        self.ranking_table = QTableWidget(0, 3)
        self.ranking_table.setHorizontalHeaderLabels(["Rank", "Name", "Max Score (dB SPL)"])
        self.ranking_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.ranking_table.setSelectionMode(QAbstractItemView.SelectionMode.NoSelection)
        self.ranking_table.verticalHeader().setVisible(False)
        self.ranking_table.setColumnWidth(0, 55)
        self.ranking_table.setColumnWidth(1, 145)
        self.ranking_table.horizontalHeader().setStretchLastSection(True)
        ranking_layout.addWidget(self.ranking_table)
        self.tv_button = QPushButton("OPEN TV DISPLAY")
        self.tv_button.clicked.connect(self.open_tv_ranking)
        ranking_layout.addWidget(self.tv_button)
        content.addWidget(ranking_box, 2)
        self.result_overlay = ResultOverlay(root)

    @staticmethod
    def _make_logo_label(parent: QWidget, size: int) -> QLabel:
        """Creates a transparent, scalable label for the supplied Monster logo."""
        logo = QLabel(parent)
        logo.setPixmap(QPixmap(str(LOGO_FILE)))
        logo.setFixedSize(size, size)
        logo.setScaledContents(True)
        logo.setStyleSheet("background: transparent; border: none;")
        return logo

    def validate_form(self) -> None:
        name_parts = self.name_input.text().strip().split()
        valid_full_name = len(name_parts) >= 2
        valid = valid_full_name and bool(EMAIL_PATTERN.match(self.email_input.text().strip()))
        self.start_button.setEnabled(valid and not self.is_measuring)

    def start_test(self) -> None:
        try:
            self.audio.start()
        except Exception as error:
            QMessageBox.critical(self, "Microphone unavailable", f"The microphone could not be opened.\n\n{error}")
            return
        self.ranking_display.show_on_tv()
        self.ranking_display.show_test(0.0, "SCREAM NOW — 5.0 SECONDS LEFT")
        self.is_measuring = True
        self.max_db = 0.0
        self.started_at = time.monotonic()
        self.start_button.setEnabled(False)
        self.name_input.setEnabled(False)
        self.email_input.setEnabled(False)
        self.status_label.setText(f"TEST ACTIVE — {self.audio.device_name} — 5.0 seconds left")
        self.measurement_timer.start()

    def update_measurement(self) -> None:
        current_db = self.audio.level()
        self.max_db = max(self.max_db, current_db)
        self.db_label.setText(f"{current_db:.1f} dB")
        self.meter.set_level(current_db)
        remaining = max(0.0, 5.0 - (time.monotonic() - self.started_at))
        self.status_label.setText(f"TEST ACTIVE — {remaining:.1f} seconds left")
        self.ranking_display.update_test(current_db, remaining)
        if remaining <= 0:
            self.finish_test()

    def finish_test(self) -> None:
        self.measurement_timer.stop()
        self.audio.stop()
        self.is_measuring = False
        score = {"name": self.name_input.text().strip(), "email": self.email_input.text().strip(), "max_db": round(self.max_db, 1)}
        scores = self.read_scores()
        scores.append(score)
        scores.sort(key=lambda entry: entry.get("max_db", 0), reverse=True)
        rank = next(index for index, entry in enumerate(scores) if entry is score) + 1
        self.write_scores(scores)
        self.populate_ranking(scores)
        self.status_label.setText(f"TEST COMPLETE — your peak: {self.max_db:.1f} dB SPL")
        self.name_input.clear()
        self.email_input.clear()
        self.name_input.setEnabled(True)
        self.email_input.setEnabled(True)
        self.validate_form()
        self.result_overlay.show_result(score["max_db"], rank == 1)
        self.ranking_display.show_result(score["max_db"], rank == 1)

    def read_scores(self) -> list[dict]:
        try:
            with HIGHSCORES_FILE.open("r", encoding="utf-8") as file:
                data = json.load(file)
            return data if isinstance(data, list) else []
        except (OSError, json.JSONDecodeError):
            return []

    def write_scores(self, scores: list[dict]) -> None:
        with HIGHSCORES_FILE.open("w", encoding="utf-8") as file:
            json.dump(scores, file, ensure_ascii=False, indent=2)

    def load_highscores(self) -> None:
        scores = self.read_scores()
        scores.sort(key=lambda entry: entry.get("max_db", 0), reverse=True)
        self.populate_ranking(scores)

    def populate_ranking(self, scores: list[dict]) -> None:
        top_ten = scores[:10]
        self.ranking_table.setRowCount(len(top_ten))
        for row, score in enumerate(top_ten):
            # Keep the full name private in JSON; show only the first name publicly.
            full_name = str(score.get("name", "Unknown")).strip()
            first_name = full_name.split()[0] if full_name else "Unknown"
            values = [str(row + 1), first_name, f"{float(score.get('max_db', 0)):.1f}"]
            for column, value in enumerate(values):
                item = QTableWidgetItem(value)
                item.setTextAlignment(Qt.AlignmentFlag.AlignCenter if column != 1 else Qt.AlignmentFlag.AlignVCenter)
                self.ranking_table.setItem(row, column, item)
        self.ranking_display.set_scores(scores)

    def open_tv_ranking(self) -> None:
        """Opens the live display full-screen on the connected TV."""
        if self.ranking_display.show_on_tv():
            self.status_label.setText("TV display opened on the second screen.")
        else:
            QMessageBox.information(
                self,
                "No second screen found",
                "Connect a TV or second monitor first. The live display will then open there full-screen.",
            )

    def schedule_daily_export(self) -> None:
        now = datetime.now()
        next_export = datetime.combine(now.date(), DAILY_EXPORT_TIME)
        if next_export <= now:
            next_export += timedelta(days=1)
        milliseconds = max(1, int((next_export - now).total_seconds() * 1000))
        self.daily_export_timer.start(milliseconds)

    def export_at_end_of_day(self) -> None:
        export_date = date.today() if TEST_EXPORT_SCHEDULE else date.today() - timedelta(days=1)
        try:
            output_file = self.export_top_ten_to_excel(export_date)
            self.send_export_email(output_file, export_date)
            self.status_label.setText(
                f"Daily Top 10 saved and emailed: {export_date.isoformat()}"
            )
        except (OSError, RuntimeError) as error:
            self.status_label.setText(f"Daily export failed: {error}")
        finally:
            self.schedule_daily_export()

    @staticmethod
    def send_export_email(output_file: Path, export_date: date) -> None:
        """Sends the finished daily Excel file through Gmail SMTP over encrypted SSL."""
        app_password = os.environ.get(SMTP_APP_PASSWORD_ENV, "").replace(" ", "")
        if not app_password:
            raise RuntimeError(
                f"Excel saved locally, but email was not sent. Set the {SMTP_APP_PASSWORD_ENV} environment variable."
            )
        message = EmailMessage()
        message["Subject"] = f"Monster Energy Scream Challenge — Top 10 — {export_date.isoformat()}"
        message["From"] = SMTP_SENDER_EMAIL
        message["To"] = EXPORT_RECIPIENT_EMAIL
        message.set_content("Attached is the daily Monster Energy Scream Challenge Top 10 export.")
        with output_file.open("rb") as attachment:
            message.add_attachment(
                attachment.read(),
                maintype="application",
                subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                filename=output_file.name,
            )
        with smtplib.SMTP_SSL(SMTP_HOST, SMTP_PORT, timeout=30) as server:
            server.login(SMTP_SENDER_EMAIL, app_password)
            server.send_message(message)

    def export_top_ten_to_excel(self, export_date: date) -> Path:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill
        except ModuleNotFoundError as error:
            raise RuntimeError("Install openpyxl with: py -m pip install openpyxl") from error
        scores = self.read_scores()
        scores.sort(key=lambda entry: entry.get("max_db", 0), reverse=True)
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Top 10"
        sheet.append(["MONSTER ENERGY SCREAM CHALLENGE"])
        sheet.merge_cells("A1:C1")
        sheet["A1"].font = Font(bold=True, size=16, color="3CD070")
        sheet["A1"].fill = PatternFill("solid", fgColor="111111")
        sheet["A1"].alignment = Alignment(horizontal="center")
        sheet.append([f"Daily Top 10 — {export_date.isoformat()}"])
        sheet.merge_cells("A2:C2")
        sheet["A2"].alignment = Alignment(horizontal="center")
        sheet.append([])
        sheet.append(["Rank", "Name", "Max Score (dB SPL)"])
        for cell in sheet[4]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="267A45")
            cell.alignment = Alignment(horizontal="center")
        for rank, score in enumerate(scores[:10], start=1):
            full_name = str(score.get("name", "Unknown")).strip()
            first_name = full_name.split()[0] if full_name else "Unknown"
            sheet.append([rank, first_name, round(float(score.get("max_db", 0)), 1)])
        sheet.column_dimensions["A"].width = 12
        sheet.column_dimensions["B"].width = 24
        sheet.column_dimensions["C"].width = 20
        for row in sheet.iter_rows(min_row=5, max_col=3):
            for cell in row:
                cell.alignment = Alignment(horizontal="center")
        EXPORT_DIR.mkdir(parents=True, exist_ok=True)
        output_file = EXPORT_DIR / f"top_10_{export_date.isoformat()}.xlsx"
        workbook.save(output_file)
        return output_file

    def closeEvent(self, event) -> None:
        self.measurement_timer.stop()
        self.daily_export_timer.stop()
        self.audio.stop()
        self.ranking_display.close()
        event.accept()

    def resizeEvent(self, event) -> None:

        super().resizeEvent(event)
        if hasattr(self, "result_overlay"):
            self.result_overlay.setGeometry(self.centralWidget().rect())


def main() -> None:
    app = QApplication(sys.argv)
    app.setStyle("Fusion")
    window = SoundboardWindow()
    window.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
